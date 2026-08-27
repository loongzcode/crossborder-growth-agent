"""Safe CSV/XLSX readers used before schema mapping."""

import csv
import io
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook  # type: ignore[import-untyped]

MAX_TABULAR_ROWS = 10_000
SUPPORTED_EXTENSIONS = frozenset({".csv", ".xlsx"})


class UnsupportedTabularFile(ValueError):
    pass


@dataclass(frozen=True, slots=True)
class TabularRows:
    rows: list[list[Any]]
    extension: str


def _decode_csv(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("CSV 编码无法识别，请使用 UTF-8 或 GB18030")


def _read_csv(content: bytes) -> list[list[Any]]:
    text = _decode_csv(content)
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    rows = [
        list(row) for row in list(csv.reader(io.StringIO(text), dialect))[: MAX_TABULAR_ROWS + 1]
    ]
    if len(rows) > MAX_TABULAR_ROWS:
        raise ValueError(f"预检最多支持 {MAX_TABULAR_ROWS} 行，请拆分文件或使用平台连接器")
    return rows


def _read_xlsx(content: bytes) -> list[list[Any]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        if sheet is None:
            return []
        rows = [
            list(row) for row in sheet.iter_rows(values_only=True, max_row=MAX_TABULAR_ROWS + 1)
        ]
        if len(rows) > MAX_TABULAR_ROWS:
            raise ValueError(f"预检最多支持 {MAX_TABULAR_ROWS} 行，请拆分文件或使用平台连接器")
        return rows
    finally:
        workbook.close()


def read_tabular(content: bytes, filename: str) -> TabularRows:
    extension = Path(filename).suffix.casefold()
    if extension not in SUPPORTED_EXTENSIONS:
        raise UnsupportedTabularFile("仅支持 .csv 和 .xlsx 广告报表")
    if not content:
        raise ValueError("上传文件为空")
    rows = _read_csv(content) if extension == ".csv" else _read_xlsx(content)
    return TabularRows(rows=rows, extension=extension)
